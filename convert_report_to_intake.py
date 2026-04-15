#!/usr/bin/env python3
"""
המרת טופס דיווח איכות מים (אתרי דלק) לטופס קליטה למערכת רשות המים.

שימוש:
    python convert_report_to_intake.py <reporting_file.xlsx> [--params <params_table.xlsx>] [--output <output.xlsx>]

אם לא מצוין קובץ פרמטרים, ברירת המחדל היא 'טבלת_פרמטרים.xlsx' באותה תיקייה.
"""

import argparse
import csv
import os
import re
import sys
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ---------------------------------------------------------------------------
# Reference table loading (CSV)
# ---------------------------------------------------------------------------

def load_csv_lookup(path: str, name_col: str, code_col: str) -> dict:
    """Load a CSV lookup table. Returns {name: code} dict."""
    mapping = {}
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            name = row[name_col].strip()
            code = int(row[code_col].strip())
            mapping[name] = code
    return mapping


# ---------------------------------------------------------------------------
# Well code memory (site+well_name -> well_code)
# ---------------------------------------------------------------------------

WELL_MEMORY_HEADERS = ['אתר', 'שם קידוח', 'קוד קידוח']


def load_well_memory(path: str) -> dict:
    """Load well code memory: {(site, well_name): well_code}."""
    memory = {}
    if not os.path.exists(path):
        return memory
    with open(path, encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                site = row['אתר'].strip()
                well_name = row['שם קידוח'].strip()
                code = int(row['קוד קידוח'].strip())
                memory[(site, well_name)] = code
            except (KeyError, ValueError):
                continue
    return memory


def save_well_memory(memory: dict, path: str):
    """Save well code memory to CSV. Creates parent directory if needed."""
    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    with open(path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=WELL_MEMORY_HEADERS)
        writer.writeheader()
        for (site, well_name), code in sorted(memory.items()):
            writer.writerow({'אתר': site, 'שם קידוח': well_name, 'קוד קידוח': code})


# ---------------------------------------------------------------------------
# Historical data loading
# ---------------------------------------------------------------------------

def load_historical_data(path: str) -> dict:
    """
    Load historical water quality data.
    Returns {(well_code, param_symbol): last_value} dict.
    For each well+param, keeps the most recent non-zero value.

    Auto-detects two formats:
    1. Intake format (headers in row 1):
       מס ש"ה, תאריך דיגום, תוצאה סופית, סמל פרמטר
    2. Historical template (headers in row 5):
       זיהוי קידוח, תאריך מדידה, ריכוז, שם פרמטר, סמן
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    # Auto-detect header row (check rows 1-10)
    # A valid header row must contain at least 2 of the expected keywords
    header_keywords = ['קידוח', 'פרמטר', 'ריכוז', 'ש"ה', 'תאריך', 'מדידה']
    header_row = None
    for candidate_row in range(1, 11):
        row_text = ''
        for col_idx in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row=candidate_row, column=col_idx).value
            if val:
                row_text += ' ' + str(val)
        matches = sum(1 for kw in header_keywords if kw in row_text)
        if matches >= 3:
            header_row = candidate_row
            break

    if header_row is None:
        print(f"  אזהרה: לא נמצאה שורת כותרות בקובץ היסטורי")
        wb.close()
        return {}

    # Read headers
    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            headers[str(val).strip()] = col_idx

    # Map columns — support both naming conventions
    col_well = (headers.get('זיהוי קידוח')
                or headers.get('מס ש"ה')
                or headers.get('מס שה'))
    col_date = (headers.get('תאריך מדידה')
                or headers.get('תאריך דיגום'))
    col_result = (headers.get('ריכוז')
                  or headers.get('תוצאה סופית'))
    col_param = (headers.get('שם פרמטר')
                 or headers.get('סמל פרמטר'))
    col_sign = headers.get('סמן') or headers.get('סימן')

    if not all([col_well, col_result, col_param]):
        print(f"  אזהרה: קובץ היסטורי חסר עמודות נדרשות. "
              f"נמצאו: {list(headers.keys())}")
        wb.close()
        return {}

    print(f"  פורמט: כותרות בשורה {header_row}, "
          f"קידוח=col {col_well}, פרמטר=col {col_param}, "
          f"ריכוז=col {col_result}")

    # Collect all records, then keep the latest per (well, param)
    records = []
    data_start = header_row + 1
    for row_idx in range(data_start, ws.max_row + 1):
        well = ws.cell(row=row_idx, column=col_well).value
        param = ws.cell(row=row_idx, column=col_param).value
        result = ws.cell(row=row_idx, column=col_result).value
        date = ws.cell(row=row_idx, column=col_date).value if col_date else None
        sign = ws.cell(row=row_idx, column=col_sign).value if col_sign else None

        if well is None or param is None or result is None:
            continue

        # Skip below-detection records (sign = "<")
        if sign and '<' in str(sign):
            continue

        if not isinstance(result, (int, float)):
            try:
                result = float(result)
            except (ValueError, TypeError):
                continue

        if result == 0:
            continue

        try:
            well_int = int(well)
        except (ValueError, TypeError):
            continue

        records.append((well_int, str(param).strip(), date, result))

    wb.close()

    # Sort by date (if available) and keep the latest value per (well, param)
    records.sort(key=lambda r: (r[0], r[1], r[2] or datetime.min))
    latest = {}
    for well, param, date, result in records:
        latest[(well, param)] = result

    return latest


def prompt_well_code(site_name: str, well_name: str, raw_value) -> int | None:
    """
    Interactively ask user for a well code.
    Returns int code or None if user chooses to skip.
    """
    print(f"\n  --- קוד קידוח חסר ---")
    print(f"  אתר: {site_name}")
    print(f"  שם קידוח: {well_name}")
    if raw_value is not None and raw_value != '-':
        print(f"  ערך בטופס: '{raw_value}'")
    while True:
        ans = input(f"  הקלד קוד קידוח (8 ספרות), או Enter לדילוג: ").strip()
        if ans == '':
            return None
        try:
            code = int(ans)
            if len(str(code)) == 8:
                return code
            else:
                print(f"  קוד קידוח צריך להיות 8 ספרות (הוקלד: {len(str(code))} ספרות)")
        except ValueError:
            print(f"  ערך לא מספרי: '{ans}'")


def fuzzy_match(name: str, known_names: list, threshold: float = 0.65) -> list:
    """
    Return list of (known_name, score) for names above threshold,
    sorted by score descending. Uses both substring and sequence matching.
    """
    results = []
    name_lower = name.lower().strip()
    for known in known_names:
        known_lower = known.lower().strip()
        # Exact match
        if name_lower == known_lower:
            results.append((known, 1.0))
            continue
        # Substring containment (either direction)
        if name_lower in known_lower or known_lower in name_lower:
            results.append((known, 0.85))
            continue
        # Sequence similarity
        score = SequenceMatcher(None, name_lower, known_lower).ratio()
        if score >= threshold:
            results.append((known, score))
    results.sort(key=lambda x: -x[1])
    return results


def validate_name_code(name: str, code, ref_lookup: dict, entity_type: str):
    """
    Validate a name/code pair against a reference lookup table.
    Returns (validated_code, errors, warnings).
    - If name matches exactly and code matches: all good.
    - If name matches exactly but code differs: warning + use ref code.
    - If name matches fuzzily: warning with suggestion.
    - If code exists in ref values but name doesn't match: warning.
    - If nothing matches: error.
    """
    errors = []
    warnings = []
    validated_code = code

    if not name:
        return validated_code, errors, warnings

    name_s = str(name).strip()

    # Exact match
    if name_s in ref_lookup:
        ref_code = ref_lookup[name_s]
        if code is not None and int(code) != ref_code:
            warnings.append(
                f"קוד {entity_type} {code} לא תואם לשם '{name_s}' "
                f"(בקובץ הייחוס: {ref_code}) — משתמש בקוד מהייחוס")
            validated_code = ref_code
        elif code is None:
            warnings.append(
                f"קוד {entity_type} חסר לשם '{name_s}' — הושלם מקובץ הייחוס: {ref_code}")
            validated_code = ref_code
        return validated_code, errors, warnings

    # Fuzzy match
    matches = fuzzy_match(name_s, list(ref_lookup.keys()))
    if matches:
        best_name, best_score = matches[0]
        best_code = ref_lookup[best_name]
        if best_score >= 0.85:
            warnings.append(
                f"שם {entity_type} '{name_s}' לא תואם בדיוק — "
                f"התאמה מקורבת: '{best_name}' (קוד {best_code})")
            if code is None:
                validated_code = best_code
        else:
            suggestions = ', '.join(f"'{n}' ({ref_lookup[n]})" for n, _ in matches[:3])
            warnings.append(
                f"שם {entity_type} '{name_s}' לא נמצא בקובץ הייחוס. "
                f"אולי: {suggestions}")
    else:
        warnings.append(f"שם {entity_type} '{name_s}' לא נמצא בקובץ הייחוס ואין התאמה מקורבת")

    return validated_code, errors, warnings


# ---------------------------------------------------------------------------
# Parameter table loading
# ---------------------------------------------------------------------------

def load_param_table(path: str) -> dict:
    """Load parameter mapping: numeric code -> symbol string."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    mapping = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        code, symbol = row[0], row[1]
        if code is not None and symbol is not None:
            mapping[int(code)] = str(symbol).strip()
    wb.close()
    return mapping



# ---------------------------------------------------------------------------
# Parse a measurement value
# ---------------------------------------------------------------------------

def parse_measurement(value):
    """
    Parse a measurement cell value.
    Returns (result_value, is_below_detection, raw_string) or None if empty/skip.
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        return (value, False, str(value))

    s = str(value).strip()
    if s == '' or s == '-':
        return None

    # Handle "<X" pattern
    m = re.match(r'^<\s*(.+)$', s)
    if m:
        return (0, True, s)

    # Try to parse as number
    try:
        return (float(s), False, s)
    except ValueError:
        return ('ERROR', False, s)


# ---------------------------------------------------------------------------
# Parse date from reporting form
# ---------------------------------------------------------------------------

def parse_date(ws):
    """Extract sampling date from row 3. Try C3 first, then B3."""
    c3 = ws.cell(row=3, column=3).value  # C3
    if isinstance(c3, datetime):
        return c3

    b3 = ws.cell(row=3, column=2).value  # B3
    if isinstance(b3, datetime):
        return b3

    # Try parsing text dates
    for val in [c3, b3]:
        if val is None:
            continue
        s = str(val).strip().replace('-', '/')
        for fmt in ['%d/%m/%Y', '%d.%m.%Y', '%d/%m/%y', '%d.%m.%y']:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue

    return None


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def convert_report(report_path: str, param_map: dict,
                   ref_labs: dict = None, ref_samplers: dict = None,
                   interactive: bool = False, well_memory: dict = None,
                   historical_data: dict = None,
                   lab_code_override: int = None,
                   sampler_code_override: int = None):
    """
    Convert a single reporting form to intake rows.
    Returns (rows, errors, warnings) where:
      - rows: list of dicts with keys A-I
      - errors: list of error strings (blocking)
      - warnings: list of warning strings (non-blocking)
    """
    errors = []
    warnings = []
    rows = []

    filename = os.path.basename(report_path)

    try:
        wb = openpyxl.load_workbook(report_path, data_only=True)
    except Exception as e:
        errors.append(f"לא ניתן לפתוח את הקובץ: {e}")
        return rows, errors, warnings

    ws = wb.active

    # --- Header extraction ---

    site_name = ws.cell(row=2, column=2).value or '?'

    # Date
    sample_date = parse_date(ws)
    if sample_date is None:
        errors.append(f"תאריך דיגום חסר או לא תקין (שורה 3)")

    # Lab code
    lab_name = ws.cell(row=4, column=2).value
    lab_code = ws.cell(row=4, column=3).value
    if lab_code is None or lab_code == '-':
        lab_code = None
    else:
        try:
            lab_code = int(lab_code)
        except (ValueError, TypeError):
            warnings.append(f"קוד מעבדה לא מספרי: '{lab_code}' (שורה 4)")
            lab_code = None

    # Validate lab against external reference
    if ref_labs and lab_name:
        lab_code, lab_errs, lab_warns = validate_name_code(
            lab_name, lab_code, ref_labs, 'מעבדה')
        errors.extend(lab_errs)
        warnings.extend(lab_warns)

    if lab_code is None:
        if lab_code_override is not None:
            lab_code = lab_code_override
            lab_name_s = str(lab_name).strip() if lab_name else '?'
            warnings.append(f"קוד מעבדה הוזן ידנית: {lab_code_override} (שם: '{lab_name_s}')")
        else:
            lab_name_s = f"'{str(lab_name).strip()}'" if lab_name else "לא ידוע"
            errors.append(f"קוד מעבדה חסר לשם {lab_name_s} ולא ניתן להשלים (שורה 4)")

    # Sampler code
    sampler_name = ws.cell(row=5, column=2).value
    sampler_code = ws.cell(row=5, column=3).value
    if sampler_code is None or sampler_code == '-':
        sampler_code = None
    else:
        try:
            sampler_code = int(sampler_code)
        except (ValueError, TypeError):
            warnings.append(f"קוד חברת דיגום לא מספרי: '{sampler_code}' (שורה 5)")
            sampler_code = None

    # Validate sampler against external reference
    if ref_samplers and sampler_name:
        sampler_code, samp_errs, samp_warns = validate_name_code(
            sampler_name, sampler_code, ref_samplers, 'חברת דיגום')
        errors.extend(samp_errs)
        warnings.extend(samp_warns)

    if sampler_code is None:
        if sampler_code_override is not None:
            sampler_code = sampler_code_override
            sampler_name_s = str(sampler_name).strip() if sampler_name else '?'
            warnings.append(f"קוד חברת דיגום הוזן ידנית: {sampler_code_override} (שם: '{sampler_name_s}')")
        else:
            sampler_name_s = f"'{str(sampler_name).strip()}'" if sampler_name else "לא ידוע"
            errors.append(f"קוד חברת דיגום חסר לשם {sampler_name_s} ולא ניתן להשלים (שורה 5)")

    # --- Identify wells (columns D onward in rows 7-8) ---
    # Scan columns starting from D. A well column has content in row 7 or 8.
    # Stop after encountering 2 consecutive empty columns (gap = end of wells).

    wells = []  # list of (col_index, well_name, well_code)
    consecutive_empty = 0
    for col_idx in range(4, 4 + 25):  # D=4, generous upper bound
        well_name = ws.cell(row=7, column=col_idx).value
        well_code = ws.cell(row=8, column=col_idx).value

        if well_name is None and well_code is None:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                break
            continue
        if well_name and str(well_name).strip() == '-':
            continue

        consecutive_empty = 0  # reset on valid column

        well_name_s = str(well_name).strip() if well_name else '?'
        needs_resolution = False
        raw_code = well_code

        # Try to parse well code as integer
        if well_code is None or well_code == '-':
            needs_resolution = True
        else:
            try:
                well_code_int = int(well_code)
                needs_resolution = False
            except (ValueError, TypeError):
                needs_resolution = True

        if needs_resolution:
            # Step 1: Check memory
            if well_memory is not None and (site_name, well_name_s) in well_memory:
                well_code_int = well_memory[(site_name, well_name_s)]
                warnings.append(
                    f"קוד קידוח לקידוח '{well_name_s}' הושלם מזיכרון: {well_code_int}")
            # Step 2: Interactive prompt
            elif interactive:
                well_code_int_or_none = prompt_well_code(site_name, well_name_s, raw_code)
                if well_code_int_or_none is not None:
                    well_code_int = well_code_int_or_none
                    # Save to memory
                    if well_memory is not None:
                        well_memory[(site_name, well_name_s)] = well_code_int
                    warnings.append(
                        f"קוד קידוח לקידוח '{well_name_s}' הוקלד ידנית: {well_code_int}")
                else:
                    errors.append(
                        f"קוד קידוח חסר לקידוח '{well_name_s}' — המשתמש דילג (עמודה {col_idx})")
                    continue
            # Step 3: No resolution possible
            else:
                if raw_code is None or raw_code == '-':
                    errors.append(f"קוד קידוח חסר לקידוח '{well_name_s}' (עמודה {col_idx})")
                else:
                    errors.append(f"קוד קידוח לא מספרי: '{raw_code}' לקידוח '{well_name_s}'")
                continue

        wells.append((col_idx, well_name_s, well_code_int))

    if not wells:
        errors.append("לא נמצאו קידוחים בטופס")

    # --- Process special header rows (10, 11) as parameters ---
    SPECIAL_ROWS = {
        10: ('WDEP', 'עומק עד פני המים'),     # row 10 -> param symbol WDEP
        11: ('OILTH', 'OIL LAYER THIKNESS'),   # row 11 -> param symbol OILTH
    }

    for row_idx, (param_symbol, expected_label) in SPECIAL_ROWS.items():
        for col_idx, well_name, well_code in wells:
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            parsed = parse_measurement(cell_value)

            if parsed is None:
                continue

            result_value, is_below_detection, raw_str = parsed

            if result_value == 'ERROR':
                errors.append(
                    f"ערך לא תקין '{raw_str}' בקידוח '{well_name}' {expected_label} (שורה {row_idx})")
                continue

            rows.append({
                'A': 5,
                'B': well_code,
                'C': sample_date,
                'D': None,
                'E': sampler_code,
                'F': None,
                'G': result_value,
                'H': param_symbol,
                'I': lab_code,
                '_source': {
                    'file': filename,
                    'well': well_name,
                    'param_code': param_symbol,
                    'row': row_idx,
                    'raw_value': raw_str,
                    'below_detection': is_below_detection,
                }
            })

    # --- Process measurement rows (row 13 onward) ---

    for row_idx in range(13, ws.max_row + 1):
        param_code_raw = ws.cell(row=row_idx, column=1).value  # A
        if param_code_raw is None:
            continue

        try:
            param_code = int(param_code_raw)
        except (ValueError, TypeError):
            # Skip non-numeric rows (could be sub-headers)
            continue

        # Map to symbol
        if param_code not in param_map:
            warnings.append(f"קוד פרמטר {param_code} (שורה {row_idx}) לא נמצא בטבלת הפרמטרים — שורות אלו ידולגו")
            continue

        param_symbol = param_map[param_code]

        # Process each well
        for col_idx, well_name, well_code in wells:
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            parsed = parse_measurement(cell_value)

            if parsed is None:
                continue  # empty cell = not measured, skip

            result_value, is_below_detection, raw_str = parsed

            if result_value == 'ERROR':
                errors.append(
                    f"ערך לא תקין '{raw_str}' בקידוח '{well_name}' פרמטר {param_code} (שורה {row_idx})")
                continue

            rows.append({
                'A': 5,                    # מקור מים - קבוע
                'B': well_code,            # מס ש"ה
                'C': sample_date,          # תאריך דיגום
                'D': None,                 # עומק דיגום - ריק
                'E': sampler_code,         # מוסד דוגם
                'F': None,                 # סימן - ריק
                'G': result_value,         # תוצאה סופית
                'H': param_symbol,         # סמל פרמטר
                'I': lab_code,             # מעבדה
                '_source': {
                    'file': filename,
                    'well': well_name,
                    'param_code': param_code,
                    'row': row_idx,
                    'raw_value': raw_str,
                    'below_detection': is_below_detection,
                }
            })

    wb.close()

    # --- Post-processing validations ---

    # Rule: ECFD must be < 100
    PARAM_MAX_VALUES = {
        'ECFD': 100,
    }
    for row_data in rows:
        param = row_data['H']
        val = row_data['G']
        if param in PARAM_MAX_VALUES and isinstance(val, (int, float)) and val != 0:
            max_val = PARAM_MAX_VALUES[param]
            if val >= max_val:
                src = row_data.get('_source', {})
                warnings.append(
                    f"ערך {param} חריג: {val} (מעל {max_val}) "
                    f"בקידוח '{src.get('well', '?')}' שורה {src.get('row', '?')}")

    # Historical anomaly detection (2 orders of magnitude)
    if historical_data is not None:
        for row_data in rows:
            well_code = row_data['B']
            param = row_data['H']
            val = row_data['G']

            if not isinstance(val, (int, float)) or val == 0:
                continue

            prev_val = historical_data.get((well_code, param))
            if prev_val is None or prev_val == 0:
                continue

            ratio = val / prev_val
            if ratio >= 100 or ratio <= 0.01:
                src = row_data.get('_source', {})
                warnings.append(
                    f"חריגה היסטורית {param} בקידוח '{src.get('well', '?')}': "
                    f"ערך נוכחי={val}, ערך קודם={prev_val} "
                    f"(יחס {ratio:.1f})")

    return rows, errors, warnings


# ---------------------------------------------------------------------------
# Write output files
# ---------------------------------------------------------------------------

def write_intake_file(rows: list, output_path: str):
    """Write intake Excel file matching the expected format."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    headers = ['מקור מים', 'מס ש"ה', 'תאריך דיגום', 'עומק דיגום',
               'מוסד דוגם', 'סימן', 'תוצאה סופית', 'סמל פרמטר', 'מעבדה']
    ws.append(headers)

    # Header formatting
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row_data in rows:
        # Format date as DD.MM.YYYY string
        date_val = row_data['C']
        if isinstance(date_val, datetime):
            date_str = date_val.strftime('%d.%m.%Y')
        else:
            date_str = date_val

        ws.append([
            row_data['A'],
            row_data['B'],
            date_str,
            row_data['D'],
            row_data['E'],
            row_data['F'],
            row_data['G'],
            row_data['H'],
            row_data['I'],
        ])

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 25)

    wb.save(output_path)
    return len(rows)


def write_error_report(all_results: list, output_path: str):
    """Write error/warning report as Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'דוח שגיאות'

    headers = ['קובץ', 'סוג', 'הודעה']
    ws.append(headers)

    header_fill_err = PatternFill('solid', fgColor='FFCCCC')
    header_fill_warn = PatternFill('solid', fgColor='FFFFCC')

    for col_idx in range(1, 4):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    row_num = 2
    has_content = False

    for filename, errors, warnings, row_count in all_results:
        for err in errors:
            ws.cell(row=row_num, column=1, value=filename)
            ws.cell(row=row_num, column=2, value='שגיאה')
            ws.cell(row=row_num, column=3, value=err)
            for c in range(1, 4):
                ws.cell(row=row_num, column=c).fill = header_fill_err
            row_num += 1
            has_content = True

        for warn in warnings:
            ws.cell(row=row_num, column=1, value=filename)
            ws.cell(row=row_num, column=2, value='אזהרה')
            ws.cell(row=row_num, column=3, value=warn)
            for c in range(1, 4):
                ws.cell(row=row_num, column=c).fill = header_fill_warn
            row_num += 1
            has_content = True

        # Summary row
        ws.cell(row=row_num, column=1, value=filename)
        ws.cell(row=row_num, column=2, value='סיכום')
        ws.cell(row=row_num, column=3,
                value=f"{row_count} שורות קליטה, {len(errors)} שגיאות, {len(warnings)} אזהרות")
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        row_num += 1
        has_content = True

    # Auto-width
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 80

    wb.save(output_path)
    return has_content


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='המרת טופסי דיווח איכות מים לטופס קליטה')
    parser.add_argument('reports', nargs='+',
                        help='קבצי דיווח (xlsx)')
    parser.add_argument('--params', default=None,
                        help='טבלת פרמטרים (xlsx). ברירת מחדל: טבלת_פרמטרים.xlsx')
    parser.add_argument('--labs', default=None,
                        help='קובץ קודי מעבדות (csv). ברירת מחדל: lab_codes.csv')
    parser.add_argument('--samplers', default=None,
                        help='קובץ קודי חברות דיגום (csv). ברירת מחדל: sampler_codes.csv')
    parser.add_argument('--output', default=None,
                        help='שם קובץ קליטה פלט. ברירת מחדל: קליטה_<timestamp>.xlsx')
    parser.add_argument('--error-report', default=None,
                        help='שם דוח שגיאות. ברירת מחדל: שגיאות_<timestamp>.xlsx')
    parser.add_argument('--interactive', action='store_true',
                        help='מצב אינטראקטיבי — שואל את המשתמש להקליד קודי קידוח חסרים')
    parser.add_argument('--well-memory', default=None,
                        help='קובץ זיכרון קודי קידוחים (csv). ברירת מחדל: well_codes_memory.csv')
    parser.add_argument('--historical', default=None,
                        help='קובץ נתונים היסטוריים (xlsx) לזיהוי חריגות')

    args = parser.parse_args()

    # Find params file
    params_path = args.params
    if params_path is None:
        for candidate in [
            os.path.join(os.path.dirname(args.reports[0]), 'טבלת_פרמטרים.xlsx'),
            'טבלת_פרמטרים.xlsx',
        ]:
            if os.path.exists(candidate):
                params_path = candidate
                break

    if params_path is None or not os.path.exists(params_path):
        print("שגיאה: לא נמצא קובץ טבלת פרמטרים. השתמש ב --params")
        sys.exit(1)

    print(f"טוען טבלת פרמטרים מ: {params_path}")
    param_map = load_param_table(params_path)
    print(f"  נטענו {len(param_map)} פרמטרים")

    # Load lab reference
    ref_labs = None
    labs_path = args.labs
    if labs_path is None:
        for candidate in [
            os.path.join(os.path.dirname(args.reports[0]), 'lab_codes.csv'),
            'lab_codes.csv',
        ]:
            if os.path.exists(candidate):
                labs_path = candidate
                break
    if labs_path and os.path.exists(labs_path):
        ref_labs = load_csv_lookup(labs_path, 'שם מעבדה', 'קוד מעבדה')
        print(f"טוען קודי מעבדות מ: {labs_path} ({len(ref_labs)} מעבדות)")
    else:
        print("אזהרה: לא נמצא קובץ קודי מעבדות — ולידציה מול רשימת הייחוס בלבד")

    # Load sampler reference
    ref_samplers = None
    samplers_path = args.samplers
    if samplers_path is None:
        for candidate in [
            os.path.join(os.path.dirname(args.reports[0]), 'sampler_codes.csv'),
            'sampler_codes.csv',
        ]:
            if os.path.exists(candidate):
                samplers_path = candidate
                break
    if samplers_path and os.path.exists(samplers_path):
        ref_samplers = load_csv_lookup(samplers_path, 'שם חברת דיגום', 'קוד חברת דיגום')
        print(f"טוען קודי חברות דיגום מ: {samplers_path} ({len(ref_samplers)} חברות)")
    else:
        print("אזהרה: לא נמצא קובץ קודי חברות דיגום — ולידציה מול רשימת הייחוס בלבד")

    # Load well memory
    well_memory = None
    well_memory_path = args.well_memory
    if well_memory_path is None:
        for candidate in [
            os.path.join(os.path.dirname(args.reports[0]), 'well_codes_memory.csv'),
            'well_codes_memory.csv',
        ]:
            well_memory_path = candidate
            break  # use this path even if file doesn't exist yet

    if args.interactive or (well_memory_path and os.path.exists(well_memory_path)):
        well_memory = load_well_memory(well_memory_path) if os.path.exists(well_memory_path) else {}
        if well_memory:
            print(f"טוען זיכרון קודי קידוחים מ: {well_memory_path} ({len(well_memory)} רשומות)")
        elif args.interactive:
            print(f"זיכרון קודי קידוחים: ריק (יישמר ב: {well_memory_path})")

    if args.interactive:
        print("מצב אינטראקטיבי: קודי קידוח חסרים יוצגו להשלמה ידנית")

    # Load historical data
    historical_data = None
    if args.historical:
        if os.path.exists(args.historical):
            print(f"טוען נתונים היסטוריים מ: {args.historical}")
            historical_data = load_historical_data(args.historical)
            print(f"  נטענו {len(historical_data)} רשומות (קידוח+פרמטר)")
        else:
            print(f"אזהרה: קובץ היסטורי לא נמצא: {args.historical}")

    # Process all reports
    all_rows = []
    all_results = []
    memory_before = dict(well_memory) if well_memory is not None else {}

    for report_path in args.reports:
        print(f"\nמעבד: {os.path.basename(report_path)}")
        if not os.path.exists(report_path):
            print(f"  שגיאה: קובץ לא נמצא")
            all_results.append((os.path.basename(report_path), ['קובץ לא נמצא'], [], 0))
            continue

        rows, errors, warnings = convert_report(
            report_path, param_map, ref_labs=ref_labs, ref_samplers=ref_samplers,
            interactive=args.interactive, well_memory=well_memory,
            historical_data=historical_data)
        all_rows.extend(rows)
        all_results.append((os.path.basename(report_path), errors, warnings, len(rows)))

        print(f"  {len(rows)} שורות קליטה")
        if errors:
            print(f"  {len(errors)} שגיאות:")
            for e in errors:
                print(f"    ✗ {e}")
        if warnings:
            print(f"  {len(warnings)} אזהרות:")
            for w in warnings:
                print(f"    ⚠ {w}")

    # Save well memory if it was modified
    if well_memory is not None and well_memory != memory_before:
        new_entries = {k: v for k, v in well_memory.items() if k not in memory_before}
        save_well_memory(well_memory, well_memory_path)
        print(f"\nזיכרון קודי קידוחים עודכן: {well_memory_path} "
              f"({len(new_entries)} רשומות חדשות, {len(well_memory)} סה\"כ)")
        for (site, wname), wcode in sorted(new_entries.items()):
            print(f"  + {site} / {wname} → {wcode}")

    # Generate output filenames
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = args.output or f'קליטה_{timestamp}.xlsx'
    error_path = args.error_report or f'שגיאות_{timestamp}.xlsx'

    # Write output
    if all_rows:
        count = write_intake_file(all_rows, output_path)
        print(f"\nנכתב קובץ קליטה: {output_path} ({count} שורות)")
    else:
        print("\nלא נוצרו שורות קליטה — בדוק את דוח השגיאות")

    write_error_report(all_results, error_path)
    print(f"נכתב דוח שגיאות: {error_path}")

    # Summary
    total_errors = sum(len(e) for _, e, _, _ in all_results)
    total_warnings = sum(len(w) for _, _, w, _ in all_results)
    print(f"\nסיכום: {len(all_rows)} שורות, {total_errors} שגיאות, {total_warnings} אזהרות")

    return 0 if total_errors == 0 else 1


if __name__ == '__main__':
    sys.exit(main())
